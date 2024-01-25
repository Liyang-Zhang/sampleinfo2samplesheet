import json
import logging

# from importlib.metadata import version as fn_version
from pathlib import Path
from typing import List

import typer
from openpyxl import load_workbook

from sampleinfo_samplesheet.cli import helper
from sampleinfo_samplesheet.config.models import ExcelConfig, HeaderEnum
from sampleinfo_samplesheet.logging.helper import (
    LoggingLevel,
    config_console_debug_logger,
)

log = logging.getLogger(__name__)
config_console_debug_logger(__name__)

APP = typer.Typer()
KUNYUAN = "kunyuan"


@APP.command(KUNYUAN)
def kunyuan(
    _: bool = typer.Option(
        None, "--version", callback=helper.version_callback, is_eager=True
    ),
    input: List[Path] = typer.Option(
        ...,
        "-i",
        "--input",
        exists=True,
        file_okay=True,
        dir_okay=False,
        help="Path to the input excel file",
    ),
    output: Path = typer.Option(
        ...,
        "-o",
        "--output",
        exists=False,
        file_okay=False,
        dir_okay=True,
        help="Directory to the output sample sheet.",
    ),
    config_file: Path = typer.Option(
        None,
        "-c",
        "--config",
        exists=True,
        file_okay=True,
        dir_okay=False,
        resolve_path=True,
        help="Path to the parameter config file.",
    ),
    verbosity: LoggingLevel = typer.Option(
        LoggingLevel.DEBUG,
        "-v",
        "--verbosity",
        envvar=("_".join([KUNYUAN, "verbosity"])).upper(),
        case_sensitive=False,
        help="logging level",
    ),
    logfile: Path = typer.Option(
        None,
        "-l",
        "--logfile",
        file_okay=True,
        dir_okay=False,
        resolve_path=True,
        envvar=("_".join([KUNYUAN, "logfile"])).upper(),
        help="path to log file",
    ),
):
    logging.basicConfig(level=verbosity.upper())
    for _name in []:
        logging.getLogger(_name).setLevel(logging.ERROR)
    rfh = helper.get_rotating_file_handler(logfile=logfile)

    if rfh:
        rfh.setLevel(verbosity.upper())
        log.addHandler(rfh)
        logging.getLogger("sampleinfo_samplesheet.cli.helper").addHandler(rfh)
        logging.getLogger("sampleinfo_samplesheet.logging.helper").addHandler(rfh)
        logging.getLogger().addHandler(rfh)

    try:
        if config_file is None:
            config = load_config(
                Path(__file__).parent.parent / "config" / "config.json"
            )
        else:
            config = load_config(config_file)

        log.debug(f"Config file: {config}")

        for f in input:
            process_single_excel(input_excel=f, config=config)

    except Exception as e:
        log.critical(f"abort due to unexpected error: {e}")


def process_single_excel(input_excel: Path, config: ExcelConfig):
    data = read_excel(file_path=input_excel, config=config)
    log.debug(f"loaded data: {data}")


def read_excel(file_path: Path, config: ExcelConfig) -> List[dict]:
    workbook = load_workbook(filename=file_path)
    sheet = workbook[config.sheet_name]
    max_row = get_maximum_rows(sheet)

    header_row = next(sheet.iter_rows(values_only=True))
    header_index = {header: idx for idx, header in enumerate(header_row)}

    data = []
    log.debug(f"the sheet headers: {header_index}")
    for row in sheet.iter_rows(min_row=2, max_row=max_row, values_only=True):
        row_data = {
            HeaderEnum(identical_header): row[header_index.get(raw_header)]
            for raw_header, identical_header in config.headers.items()
        }
        data.append(row_data)
    return data


def load_config(config_file: Path) -> ExcelConfig:
    with open(config_file, "r") as file:
        config_data = json.load(file)
    config_data["headers"] = {
        k: HeaderEnum[v] for k, v in config_data["headers"].items()
    }
    return ExcelConfig(**config_data)


def get_maximum_rows(sheet_object):
    rows = 0
    for _, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows


@APP.callback()
def callback():
    """
    Convert a sequencing sample info excel to a sample sheet for bcl2fastq
    """


def main():
    APP()


if __name__ == "__main__":
    main()
